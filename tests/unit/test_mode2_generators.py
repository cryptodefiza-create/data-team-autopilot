"""Phase 2 tests: XLSX/CSV generators, provenance, interpretation, sanitizer."""

import io
from datetime import datetime, timezone

import openpyxl

from data_autopilot.services.mode1.csv_generator import CSVGenerator
from data_autopilot.services.mode1.interpretation import (
    InterpretationEngine,
    sanitize_prompt,
)
from data_autopilot.services.mode1.models import (
    DataRequest,
    Entity,
    Intent,
    Interpretation,
    Provenance,
)
from data_autopilot.services.mode1.xlsx_generator import XLSXGenerator

_SAMPLE_HOLDERS = [{"wallet": f"addr_{i}", "amount": 1000 - i * 10} for i in range(100)]


def test_xlsx_generation() -> None:
    """2.1: XLSX with Data + Metadata sheets, provenance on Metadata sheet."""
    gen = XLSXGenerator()
    prov = Provenance(
        source="Helius RPC (Solana mainnet)",
        timestamp=datetime(2025, 6, 1, 12, 0, tzinfo=timezone.utc),
        chain="solana",
        record_count=100,
    )
    xlsx_bytes = gen.generate(_SAMPLE_HOLDERS, prov)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "Data" in wb.sheetnames
    assert "Metadata" in wb.sheetnames

    ws_data = wb["Data"]
    assert ws_data.cell(1, 1).value == "wallet"
    assert ws_data.cell(1, 2).value == "amount"
    assert ws_data.cell(2, 1).value == "addr_0"
    assert ws_data.max_row == 101  # header + 100 rows

    ws_meta = wb["Metadata"]
    assert ws_meta.cell(1, 2).value == "Helius RPC (Solana mainnet)"
    assert ws_meta.cell(3, 2).value == "solana"
    assert ws_meta.cell(5, 2).value == "100"


def test_csv_generation() -> None:
    """2.2: CSV with provenance comment header."""
    gen = CSVGenerator()
    prov = Provenance(
        source="CoinGecko API",
        timestamp=datetime(2025, 6, 1, 12, 0, tzinfo=timezone.utc),
        chain="cross_chain",
        record_count=100,
    )
    csv_str = gen.generate(_SAMPLE_HOLDERS, prov)
    assert isinstance(csv_str, str)

    lines = csv_str.strip().split("\n")
    # First lines should be provenance comments
    assert lines[0].startswith("# Source: CoinGecko API")
    assert lines[1].startswith("# Queried:")
    assert lines[4].startswith("# Truncated:")
    # Then CSV header
    assert "wallet" in lines[5]
    assert "amount" in lines[5]


def test_provenance_footer() -> None:
    """2.3: Provenance footer has source, timestamp, record count."""
    prov = Provenance(
        source="Helius RPC (Solana mainnet)",
        timestamp=datetime(2025, 6, 1, 12, 0, tzinfo=timezone.utc),
        chain="solana",
        record_count=5000,
        truncated=True,
        sampling_note="Showing top 100 by balance",
    )
    footer = prov.format_footer()
    assert "Helius RPC (Solana mainnet)" in footer
    assert "Jun 01, 2025 12:00 UTC" in footer
    assert "5,000" in footer
    assert "(truncated)" in footer
    assert "Showing top 100 by balance" in footer


def test_interpretation_holders() -> None:
    """2.4: Interpretation of holder data produces observations, no advice."""
    engine = InterpretationEngine()
    holders = [{"wallet": f"w{i}", "amount": 10000 / (i + 1)} for i in range(5000)]
    request = DataRequest(
        raw_message="Show holders of $BONK",
        intent=Intent.SNAPSHOT,
        entity=Entity.TOKEN_HOLDERS,
        token="BONK",
    )
    result = engine.interpret(holders, request)
    assert isinstance(result, Interpretation)
    assert len(result.text) > 0
    assert result.disclaimer
    assert "advice" in result.disclaimer.lower()
    # Stats should have concentration data
    assert "top_holders_pct" in result.stats


def test_sanitizer_removes_api_keys() -> None:
    """2.5: sanitize_prompt removes API key patterns from LLM prompt."""
    text = (
        "Query with key sk_live_abc123def456ghi789 "
        "and shopify token shpat_xyz123 "
        "and hex key aabbccddee112233445566778899aabbccddeeff0011223344556677889900aa"
    )
    cleaned = sanitize_prompt(text)
    assert "sk_live_" not in cleaned
    assert "shpat_" not in cleaned
    assert "aabbccddee" not in cleaned
    assert "[REDACTED]" in cleaned


def test_tier_limit_free_interpretation() -> None:
    """2.16: Free tier gets interpretation on top 50 only, full data still available."""
    engine = InterpretationEngine()
    data = [{"wallet": f"w{i}", "amount": 100 - i} for i in range(5000)]
    request = DataRequest(
        raw_message="Show holders",
        intent=Intent.SNAPSHOT,
        entity=Entity.TOKEN_HOLDERS,
        token="TEST",
    )
    # max_rows_for_stats=50 means only first 50 rows feed into stats
    result = engine.interpret(data, request, max_rows_for_stats=50)
    assert result.stats["total_records"] == 50  # stats computed on 50 rows
