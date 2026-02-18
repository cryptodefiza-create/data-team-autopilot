"""Phase 2 tests: Conversation memory, follow-ups, exports, data transformer."""

import io
from unittest.mock import MagicMock, patch

import openpyxl

from data_autopilot.services.mode1.conversation_memory import ConversationMemory
from data_autopilot.services.mode1.data_transformer import DataTransformer
from data_autopilot.services.mode1.file_upload import FileUploadHandler
from data_autopilot.services.mode1.live_fetcher import LiveFetcher
from data_autopilot.services.mode1.models import (
    Chain,
    DataRequest,
    Entity,
    Intent,
    OutputFormat,
    ProviderResult,
    RoutingDecision,
    RoutingMode,
)
from data_autopilot.services.mode1.platform_keys import PlatformKeyManager
from data_autopilot.services.mode1.request_parser import RequestParser


def _make_fetcher_with_mock(records: list[dict]) -> LiveFetcher:
    mock_provider = MagicMock()
    mock_provider.fetch.return_value = ProviderResult(
        provider="test",
        method="test_method",
        records=records,
        total_available=len(records),
    )
    return LiveFetcher(
        providers={"coingecko": mock_provider},
        key_manager=PlatformKeyManager(),
        parser=RequestParser(),
        tier="pro",
    )


def test_follow_up_query() -> None:
    """2.9: Follow-up with filter applies to previous result."""
    fetcher = _make_fetcher_with_mock(
        [{"wallet": f"w{i}", "balance": 1000 + i * 100} for i in range(20)]
    )

    # First query stores data
    with patch.object(fetcher._router, "route") as mock_route:
        mock_route.return_value = RoutingDecision(
            mode=RoutingMode.PUBLIC_API,
            confidence=0.9,
            provider_name="coingecko",
            method_name="get_price",
        )
        fetcher.handle("Show holders of $PEPE", session_id="s1")

    # Follow-up filter
    result = fetcher.handle(
        "Now filter to balance > 1500",
        session_id="s1",
        filters={"balance_min": 1500},
    )
    assert result["response_type"] == "blockchain_result"
    # All records should have balance >= 1500
    for rec in result["data"]["records"]:
        assert rec["balance"] >= 1500


def test_follow_up_export() -> None:
    """2.10: Export previous filtered data as XLSX."""
    fetcher = _make_fetcher_with_mock(
        [{"wallet": f"w{i}", "balance": 500 + i * 100} for i in range(10)]
    )

    # First query
    with patch.object(fetcher._router, "route") as mock_route:
        mock_route.return_value = RoutingDecision(
            mode=RoutingMode.PUBLIC_API,
            confidence=0.9,
            provider_name="coingecko",
            method_name="get_price",
        )
        fetcher.handle("Show holders of $BONK", session_id="s2")

    # Filter first
    fetcher.handle(
        "filter to balance > 1000",
        session_id="s2",
        filters={"balance_min": 1000},
    )

    # Export filtered data
    result = fetcher.handle("Export that as spreadsheet", session_id="s2")
    assert result["response_type"] == "blockchain_export"
    assert result["data"]["format"] == "xlsx"
    assert isinstance(result["data"]["content_bytes"], bytes)

    # Verify it's a valid XLSX
    wb = openpyxl.load_workbook(io.BytesIO(result["data"]["content_bytes"]))
    assert "Data" in wb.sheetnames
    ws = wb["Data"]
    # All exported records should have balance > 1000
    for row_idx in range(2, ws.max_row + 1):
        balance = ws.cell(row_idx, 2).value
        assert balance >= 1000


def test_transformer_filter() -> None:
    """2.11: DataTransformer filter with _min suffix."""
    transformer = DataTransformer()
    data = [{"name": "a", "balance": 500}, {"name": "b", "balance": 1500}, {"name": "c", "balance": 2000}]
    result = transformer.filter(data, {"balance_min": 1000})
    assert len(result) == 2
    assert all(r["balance"] >= 1000 for r in result)


def test_transformer_aggregate() -> None:
    """2.12: DataTransformer aggregate groups and counts."""
    transformer = DataTransformer()
    data = [
        {"tier": "whale", "balance": 10000},
        {"tier": "whale", "balance": 20000},
        {"tier": "retail", "balance": 100},
        {"tier": "retail", "balance": 200},
        {"tier": "retail", "balance": 50},
    ]
    result = transformer.aggregate(data, group_by="tier", metrics=["balance"])
    assert len(result) == 2
    whale = next(r for r in result if r["tier"] == "whale")
    assert whale["count"] == 2
    assert whale["balance_sum"] == 30000
    retail = next(r for r in result if r["tier"] == "retail")
    assert retail["count"] == 3


def test_transformer_computed_column() -> None:
    """2.13: DataTransformer adds % of supply column correctly."""
    transformer = DataTransformer()
    data = [
        {"wallet": "a", "balance": 500},
        {"wallet": "b", "balance": 300},
        {"wallet": "c", "balance": 200},
    ]
    result = transformer.add_computed_columns(
        data, [{"name": "pct_supply", "field": "balance", "operation": "pct_of_total"}]
    )
    assert result[0]["pct_supply"] == 50.0
    assert result[1]["pct_supply"] == 30.0
    assert result[2]["pct_supply"] == 20.0


def test_file_upload_csv() -> None:
    """2.14: CSV file upload parses all rows correctly."""
    handler = FileUploadHandler()
    csv_content = b"wallet,balance\naddr_1,1000\naddr_2,2000\naddr_3,3000"
    result = handler.process(csv_content, "text/csv", "wallets.csv")
    assert result.record_count == 3
    assert result.records[0]["wallet"] == "addr_1"
    assert result.records[2]["balance"] == "3000"


def test_file_upload_xlsx() -> None:
    """2.15: XLSX file upload parses records from first sheet."""
    # Create a small XLSX in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["wallet", "balance"])
    ws.append(["addr_1", 1000])
    ws.append(["addr_2", 2000])
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    handler = FileUploadHandler()
    result = handler.process(
        xlsx_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "data.xlsx",
    )
    assert result.record_count == 2
    assert result.records[0]["wallet"] == "addr_1"
    assert result.records[1]["balance"] == 2000
