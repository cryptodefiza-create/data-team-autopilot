from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from data_autopilot.services.mode1.models import Provenance


class XLSXGenerator:
    def generate(self, data: list[dict[str, Any]], metadata: Provenance) -> bytes:
        wb = openpyxl.Workbook()

        # Data sheet
        ws = wb.active
        ws.title = "Data"
        if data:
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h, "") for h in headers])
            # Auto-size columns
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(header))
                for row in data[:100]:
                    val_len = len(str(row.get(header, "")))
                    if val_len > max_len:
                        max_len = val_len
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Metadata sheet
        ws_meta = wb.create_sheet("Metadata")
        ws_meta.append(["Source", metadata.source])
        ws_meta.append(["Queried", metadata.timestamp.isoformat()])
        ws_meta.append(["Chain", metadata.chain or ""])
        ws_meta.append(["Parameters", str(metadata.params)])
        ws_meta.append(["Records", str(metadata.record_count)])
        ws_meta.append(["Truncated", str(metadata.truncated)])
        ws_meta.append(["Filters", str(metadata.filters)])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
